import geopandas as gpd
import json
import os

### The "produceJoin" method is derived from the "spatial_join_geojson" script
### https://github.com/wendybw/optimizing-Seattle-PDIS/blob/main/spatial_join_geojson.py
### This script is edited to allow for automation

## This script does not need to be run in QGIS
# Replicates the "Join attributes by location" tool in QGIS
def produceJoin(dir_walkshed, walkshed_name, dir_amenity, amenity_name, dir_output, output_name):
    # Load your buffered walkshed GeoJSON file
    buffered_walksheds = gpd.read_file(dir_walkshed + "\\" + walkshed_name)

    # Load your amenities GeoJSON file
    gdf_amenities = gpd.read_file(dir_amenity + "\\" + amenity_name)

    # Ensure the CRS for amenities matches that of the buffered walksheds, reproject if necessary
    if gdf_amenities.crs != buffered_walksheds.crs:
        gdf_amenities = gdf_amenities.to_crs(buffered_walksheds.crs)

    # Perform spatial joins for reachable and unreachable amenities
    reachable_amenities = gpd.sjoin(gdf_amenities, buffered_walksheds, how="inner", predicate='intersects')
    unreachable_amenities = gpd.sjoin(gdf_amenities, buffered_walksheds, how="left", predicate='intersects')
    unreachable_amenities = unreachable_amenities[unreachable_amenities.index_right.isnull()]

    # Cleanup before exporting
    reachable_amenities.drop(columns=['index_right'], inplace=True)
    unreachable_amenities.drop(columns=['index_right'], inplace=True)

    # Export results to GeoJSON
    if not unreachable_amenities.empty:
        unreachable_file_name = dir_output + "\\Unreachable " + output_name + ".geojson"
        unreachable_amenities.to_file(unreachable_file_name, driver='GeoJSON')
    else:
        print("No unreachable amenities found.")

    if not reachable_amenities.empty:
        reachable_file_name = dir_output + "\Reachable " + output_name + ".geojson"
        reachable_amenities.to_file(reachable_file_name, driver='GeoJSON')
        print("Reachable and unreachable amenities exported successfully.")
        # Return reachable file name if reachable amenities
        return reachable_file_name
        
    else:
        print("No reachable amenities found.")
        print("Reachable and unreachable amenities exported successfully.")
        # Return unreachable file name if reachable amenities do not exist
        return unreachable_file_name

   
   
# Outputs number of joined features (number of amenities a walkshed intersects) into xlsx file
def insertXLSX(workbook, file_path, location, profile, amenity, dict_location, dict_profile):
    import openpyxl

    # Load xlsx file
    wb = openpyxl.load_workbook(workbook)
    ws = wb.active

    # Determine row in xlsx to insert data based on housing location and mobility profile of walkshed
    loc_index = list(dict_location.keys()).index(location)
    prof_index = dict_profile[profile]
    row = (loc_index * 6) + 1 + prof_index

    # Insert housing location and mobility profile in xlsx sheet
    new_data = [location, profile]
    for col, value in enumerate(new_data, start=1):
        ws.cell(column=col, row=row, value=value)
        #print("column: " + str(col) + "  row: " + str(row) + "  value: " + str(value))


    # Read geojson file in 'utf-8' encoding
    with open(file_path, 'r', encoding='utf-8') as json_data:
        data = json.load(json_data)
    # Determine how many amenities (features) are intersected by the walkshed
    val = len(data['features'])

    # Number of amenities intersected by walkshed is 0 if the parameterized joined file is the unreachable amenities
    # Check if reachable amenities do not exist
    file_name = os.path.basename(file_path)
    file_name_tokens = file_name.split(" ")
    if file_name_tokens[0] == "Unreachable": # First word in basename of parameterized file should be "Reachable" or "Unreachable"
        val = 0

    # Insert number of reachable amenities into xlsx file
    for index, column in enumerate(ws.iter_cols(min_col=3, max_col=7), start=3):
        if column[0].value == amenity:
            ws.cell(column=index, row=row, value=val)
            print("column: " + str(index) + "  row: " + str(row) + "  value: " + str(val))
            

    wb.save(workbook)



## Change file and folder locations below
## All input file names should begin with "Buffer " - naming convention directly follows from "CreateBuffers" script
# File location of xlsx sheet to store number of amenities that intersect walkshed
workbook = r"C:\QGIS Projects\Walkshed Sim\Amenity Reachability Comparison Table.xlsx"

# Assume folders only have walksheds and only have amenities
dir_walkshed = r"C:\QGIS Projects\Walkshed Sim\Buffered Walksheds"
dir_amenity = r"C:\QGIS Projects\Walkshed Sim\Buffered Amenities"

dir_output = r"C:\QGIS Projects\Walkshed Sim\Final Joined Files"

## Change dictionaries to match user preferences
# Housing region/location of each walkshed : location of each amenity
dict_location = { 
    "Argyle": "South",
    "BerthaPitts": "South",
    "Burbridge": "Burbridge",
    "E.Republican": "South",
    "MaryPilgrim": "North",
    "NorthStar": "North",
    "SacredMed": "North",
    "SalmonberryL": "South"
}

# Assign order of mobility profiles to be inserted into the xlsx file
dict_profile = {
    "Control": 1,
    "Walking": 2,
    "Cane": 3,
    "Powered": 4,
    "Manual": 5
}


# Loop through every walkshed in walkshed directory
for walkshed in os.listdir(dir_walkshed):
    walkshed_name, walkshed_ext = os.path.splitext(walkshed)

    # Check if walkshed file is a geojson
    if walkshed_ext == ".geojson":
        walkshed_name_tokens = walkshed_name.split(" ")
        # Determine housing location of walkshed
        location_w = walkshed_name_tokens[1]

        # Loop through every amenity in amenity directory corresponding to walkshed
        for amenity in os.listdir(dir_amenity):
            amenity_name, amenity_ext = os.path.splitext(amenity)

            # Check if amenity file is a geojson
            if amenity_ext == ".geojson":
                amenity_name_tokens = amenity_name.split(" ")

                # Check if amenity location corresponds to walkshed location
                location_a = amenity_name_tokens[2]
                if location_a == dict_location[location_w]:
                    # Create base name for joined file
                    output_name = " ".join(walkshed_name_tokens[1:]) + " " + " ".join(amenity_name_tokens[3:])

                    # Produce joined walkshed and amenity file, stored in variable "file_name"
                    file_name = produceJoin(dir_walkshed, walkshed, dir_amenity, amenity, dir_output, output_name)

                    # Insert number of reachable amenities into xlsx file
                    insertXLSX(workbook, file_name, location_w, walkshed_name_tokens[3], " ".join(amenity_name_tokens[3:]),
                            dict_location, dict_profile)
                    
            else:
                print("Error: " + amenity + " is not a geojson file")
    else:
        print("Error: " + walkshed + " is not a geojson file")